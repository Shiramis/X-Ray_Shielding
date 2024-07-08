from tkinter import *
from tkinter import ttk

class dqprimsec():
    def primary(self):  # Primary Barrier calculations
        import math
        self.p = float(self.entry1.get())
        self.ksec = float(self.entry4.get())
        from openpyxl import load_workbook  # Opening excel Data
        wb = load_workbook('Data Shielding.xlsx')
        # ========================primary αβγ========================
        ws = wb['prim abc']
        for i in range(3, 38):
            if self.om1.get() == ws['A' + str(i)].value:
                self.alp = ws['B' + str(i)].value
                self.blp = ws['C' + str(i)].value
                self.clp = float(ws['D' + str(i)].value)
                self.acp = ws['E' + str(i)].value
                self.bcp = ws['F' + str(i)].value
                self.ccp = float(ws['G' + str(i)].value)
                self.agp = ws['H' + str(i)].value
                self.bgp = ws['I' + str(i)].value
                self.cgp = float(ws['J' + str(i)].value)
                self.asp = ws['K' + str(i)].value
                self.bsp = ws['L' + str(i)].value
                self.csp = float(ws['M' + str(i)].value)
                self.appr = ws['N' + str(i)].value
                self.bpp = ws['O' + str(i)].value
                self.cpp = float(ws['P' + str(i)].value)
                self.awp = ws['Q' + str(i)].value
                self.bwp = ws['R' + str(i)].value
                self.cwp = float(ws['S' + str(i)].value)
        #===============Workload=======
        if self.vrawork.get() == 2:
            self.n = int(self.entry3.get())
        elif self.vrawork.get() == 1:
            ws = wb['Workload']
            for x in range(2, 12):
                if self.var_om1.get() == ws['A' + str(x)].value:
                    self.n = float(self.worentry.get()) / (ws['B' + str(x)].value)
        # ================= Occupancy factor Τ=================
        if self.vraoccup.get() == 1:
            self.t = float(self.occupentry.get())
        elif self.vraoccup.get() == 2:
            ws = wb['Occupancy Factor ( T )']
            for i in range(2, 31):
                if self.var_om2.get() == ws['A' + str(i)].value:
                    self.t = float(ws['B' + str(i)].value)
        # ====================K1" air kerma"================
        ws = wb['unshielding prim']
        for i in range(2, 5):
            if self.var_om1.get() == ws['A' + str(i)].value:
                self.K1 = float(ws['C' + str(i)].value)
            else:
                self.K1 = 1
        # =====================Use Factor=====================
        self.Us = float(self.use_ent.get())
        # ===============Preshielding=====================
        if self.preshvar.get() == 1:
            ws = wb["Equiv. thickness of prim pres"]
            if self.radiob_pre.get() == 1:
                self.xlead = float(ws['B' + str(3)].value)
                self.xconc = float(ws['C' + str(3)].value)
                self.xsteel = float(ws['D' + str(3)].value)
            elif self.radiob_pre.get() == 2:
                self.xlead = float(ws['B' + str(4)].value)
                self.xconc = float(ws['C' + str(4)].value)
                self.xsteel = float(ws['D' + str(4)].value)
        else:
            self.xlead = 0
            self.xconc = 0
            self.xsteel = 0
        # ===================== Bprimary======================================================

        self.Bpri = (self.ksec * self.p ** 2) / (self.n *self.Us* self.t * self.K1)
        # ========================== Material selection =================================
        if self.om3.get() == "Lead":
            self.xlpbar = (1 / (self.alp * self.clp)) * math.log(
                (self.Bpri ** (-self.clp) + self.blp / self.alp) / (1 + self.blp / self.alp))-self.xlead
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Lead shielding is " + str(round(self.xlpbar, 2)) + ' mm')
        elif self.om3.get() == "Concrete":
            self.xcpbar = (1 / (self.acp * self.ccp)) * math.log(
                (self.Bpri ** (-self.ccp) + self.bcp / self.acp) / (1 + self.bcp / self.acp))-self.xconc
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Concrete shielding is " + str(round(self.xcpbar, 2)) + ' mm')
        elif self.om3.get() == "Gypsum Wallboard":
            self.xgpbar = (1 / (self.agp * self.cgp)) * math.log(
                (self.Bpri ** (-self.cgp) + self.bgp / self.agp) / (1 + self.bgp / self.agp))
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Gypsum Wallboard shielding is " + str(round(self.xgpbar, 2)) + ' mm')
        elif self.om3.get() == "Steel":
            self.xspbar = (1 / (self.asp * self.csp)) * math.log(
                (self.Bpri ** (-self.csp) + self.bsp / self.asp) / (1 + self.bsp / self.asp))-self.xsteel
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Steel shielding is " + str(round(self.xspbar, 2)) + ' mm')
        elif self.om3.get() == "Plate Glass":
            self.xppbar = (1 / (self.appr * self.cpp)) * math.log(
                (self.Bpri ** (-self.cpp) + self.bpp / self.appr) / (1 + self.bpp / self.appr))
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Plate Glass shielding is " + str(round(self.xppbar, 2)) + ' mm')
        elif self.om3.get() == "Wood":
            self.xwpbar = (1 / (self.awp * self.cwp)) * math.log(
                (self.Bpri ** (-self.cwp) + self.bwp / self.awp) / (1 + self.bwp / self.awp))
            self.f6 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Wood shielding is " + str(round(self.xwpbar, 2)) + ' mm')
        self.f6.grid(row=1, column=0, pady=0, padx=20, sticky="w")

    def secondary(self):  # Secondary barrier calculations
        import math
        self.s = float(self.entry2.get())
        self.ksec = float(self.entry4.get())
        from openpyxl import load_workbook  # Opening excel Data
        wb = load_workbook('Data Shielding.xlsx')
        ws = wb['sec abc']
        # ====================secondary αβγ============================
        for i in range(3, 18):
            if self.om1.get() == ws['A' + str(i)].value:
                self.als = ws['B' + str(i)].value
                self.bls = ws['C' + str(i)].value
                self.cls = float(ws['D' + str(i)].value)
                self.acs = ws['E' + str(i)].value
                self.bcs = ws['F' + str(i)].value
                self.ccs = float(ws['G' + str(i)].value)
                self.ags = ws['H' + str(i)].value
                self.bgs = ws['I' + str(i)].value
                self.cgs = float(ws['J' + str(i)].value)
                self.ass = ws['K' + str(i)].value
                self.bss = ws['L' + str(i)].value
                self.css = float(ws['M' + str(i)].value)
                self.aps = ws['N' + str(i)].value
                self.bps = ws['O' + str(i)].value
                self.cps = float(ws['P' + str(i)].value)
                self.aws = ws['Q' + str(i)].value
                self.bws = ws['R' + str(i)].value
                self.cws = float(ws['S' + str(i)].value)
        # ========Workload===========
        if self.vrawork.get() == 2:
            self.n = int(self.entry3.get())
        elif self.vrawork.get() == 1:
            ws = wb['Workload']
            for x in range(3, 12):
                if self.var_om1.get() == ws['A' + str(x)].value:
                    self.n = float(self.worentry.get()) / (ws['B' + str(x)].value)
        # ================= Occupancy factor Τ=================
        if self.vraoccup.get() == 1:
            self.t = float(self.occupentry.get())
        elif self.vraoccup.get() == 2:
            ws = wb['Occupancy Factor ( T )']
            for i in range(2, 31):
                if self.var_om2.get() == ws['A' + str(i)].value:
                    self.t = float(ws['B' + str(i)].value)
        # ====================K1"secondary air kerma"================
        if self.leakvar.get() == 1:
            ws = wb["Uns Air Kerma"]
            if self.radiob_leak.get() == 1:
                for i in range(2, 10):
                    if self.var_om1.get() == ws['A' + str(i)].value:
                        self.K1 = float(ws['G' + str(i)].value)
            if self.radiob_leak.get() == 2:
                for i in range(2, 10):
                    if self.var_om1.get() == ws['A' + str(i)].value:
                        self.K1 = float(ws['I' + str(i)].value)
        else:
            self.K1 = 1
        # ===================== Bsec===============================================================
        self.Bsec = (self.ksec * self.s ** 2) / (self.n * self.t *self.K1)
        # ========================== Material selection =================================
        if self.om3.get() == "Lead":
            self.xlsbar = (1 / (self.als * self.cls)) * math.log(
                (self.Bsec ** (-self.cls) + self.bls / self.als) / (1 + self.bls / self.als))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Lead shielding is " + str(round(self.xlsbar, 2)) + ' mm')
        elif self.om3.get() == "Concrete":
            self.xcsbar = (1 / (self.acs * self.ccs)) * math.log(
                (self.Bsec ** (-self.ccs) + self.bcs / self.acs) / (1 + self.bcs / self.acs))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Concrete shielding is " + str(round(self.xcsbar, 2)) + ' mm')
        elif self.om3.get() == "Gypsum Wallboard":
            self.xgsbar = (1 / (self.ags * self.cgs)) * math.log(
                (self.Bsec ** (-self.cgs) + self.bgs / self.ags) / (1 + self.bgs / self.ags))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Gypsum Wallboard shielding is " + str(round(self.xgsbar, 2)) + ' mm')
        elif self.om3.get() == "Steel":
            self.xssbar = (1 / (self.ass * self.css)) * math.log(
                (self.Bsec ** (-self.css) + self.bss / self.ass) / (1 + self.bss / self.ass))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Steel shielding is " + str(round(self.xssbar, 2)) + ' mm')
        elif self.om3.get() == "Plate Glass":
            self.xpsbar = (1 / (self.aps * self.cps)) * math.log(
                (self.Bsec ** (-self.cps) + self.bps / self.aps) / (1 + self.bps / self.aps))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Plate Glass shielding is " + str(round(self.xpsbar, 2)) + ' mm')
        elif self.om3.get() == "Wood":
            self.xwsbar = (1 / (self.aws * self.cws)) * math.log(
                (self.Bsec ** (-self.cws) + self.bws / self.aws) / (1 + self.bws / self.aws))
            self.f7 = ttk.Label(master=self.frame_right, style="BL.TLabel",
                                text="Thickness of Wood shielding is " + str(round(self.xwsbar, 2)) + ' mm')
        self.f7.grid(row=3, column=0, pady=0, padx=20, sticky="w")