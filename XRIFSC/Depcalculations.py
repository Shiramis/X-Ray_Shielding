from tkinter import *
from tkinter import ttk

import pandas as pd
import math

class departprimsec():

    def choosetype(self,e,nr,ep,t):
        if self.d["vselroom "+str(t)].get() == "CT Room":
            self.depCTcal(e, nr, ep, t)
        else:
            if self.d["radiob_w "+str(e)+nr].get()== 1:
                self.depprimary(e,nr,ep,t)
            elif self.d["radiob_w "+str(e)+nr].get()== 2:
                self.depsecondary(e,nr,ep,t)

    def depprimary(self,e,nr,ep,t):  # Primary Barrier calculations
        if self.d["titleresul " + str(e)+nr] is not None:
            for o in range(1, 7):
                if self.res["resmat " + str(o) + str(e)] is not None:
                    self.res["resmat " + str(o)+str(e)].destroy()
        else:
            for o in range(1, 7):
                self.res["resmat {0}".format(str(o)) + str(e)] = None

        for o in range(1,self.d["vnumbmat " +str(e)+nr].get()+1):
            self.p = float(self.d["entryd " + str(e)+nr].get())
            if self.d["vselroom {0}".format(str(t))].get()=="From X-Ray room":
                self.ksec = float(self.d["dikeent " + str(t)].get())
            else:
                self.ksec = float(self.d["dikeent " + str(e) + nr].get())

            from openpyxl import load_workbook  # Opening excel Data
            wb = load_workbook('Data Shielding.xlsx')
            #========Workload===========
            if self.d["vselroom "+str(t)].get() == "X-Ray room":
                if self.d["vrawork " + str(t)].get() == 2:
                    self.n = int(self.d["numpapwe " + str(t)].get())
                elif self.d["vrawork " + str(t)].get() == 1:
                    ws = wb['Workload']
                    for x in range(1, 13):
                        if self.d["vsexroom " + str(t)].get() == ws['A' + str(x)].value:
                            self.n = float(self.d["worentry " + str(t)].get())/ (ws['B' + str(x)].value)
                    if self.d["vselxray "+ str(t)].get() == 2:
                        self.n = float(self.d["worentry " + str(t)].get()) /2.5
                # ========================primary αβγ========================
                ws = wb['prim abc']
                for x in range(2, 39):
                    if self.d["vsexroom " + str(t)].get() == ws['A' + str(x)].value:
                        self.alp = ws['B' + str(x)].value
                        self.blp = ws['C' + str(x)].value
                        self.clp = float(ws['D' + str(x)].value)
                        self.acp = ws['E' + str(x)].value
                        self.bcp = ws['F' + str(x)].value
                        self.ccp = float(ws['G' + str(x)].value)
                        self.agp = ws['H' + str(x)].value
                        self.bgp = ws['I' + str(x)].value
                        self.cgp = float(ws['J' + str(x)].value)
                        self.asp = ws['K' + str(x)].value
                        self.bsp = ws['L' + str(x)].value
                        self.csp = float(ws['M' + str(x)].value)
                        self.appr = ws['N' + str(x)].value
                        self.bpp = ws['O' + str(x)].value
                        self.cpp = float(ws['P' + str(x)].value)
                        self.awp = ws['Q' + str(x)].value
                        self.bwp = ws['R' + str(x)].value
                        self.cwp = float(ws['S' + str(x)].value)
            else:
                if self.d["vrawork " + str(e)+nr].get() == 2:
                    self.n = int(self.d["numpapwe " + str(e)+nr].get())
                elif self.d["vrawork " + str(e)+nr].get() == 1:
                    ws = wb['Workload']
                    for x in range(2, 13):
                        if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(x)].value:
                            self.n = float(self.d["worentry " + str(e)+nr].get())/ (ws['B' + str(x)].value)
                    if self.d["vselxray "+ str(e)+nr].get() == 2:
                        self.n = float(self.d["worentry " + str(e)+nr].get()) /2.5
                # ========================primary αβγ========================
                ws = wb['prim abc']
                for x in range(3, 39):
                    if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(x)].value:
                        self.alp = ws['B' + str(x)].value
                        self.blp = ws['C' + str(x)].value
                        self.clp = float(ws['D' + str(x)].value)
                        self.acp = ws['E' + str(x)].value
                        self.bcp = ws['F' + str(x)].value
                        self.ccp = float(ws['G' + str(x)].value)
                        self.agp = ws['H' + str(x)].value
                        self.bgp = ws['I' + str(x)].value
                        self.cgp = float(ws['J' + str(x)].value)
                        self.asp = ws['K' + str(x)].value
                        self.bsp = ws['L' + str(x)].value
                        self.csp = float(ws['M' + str(x)].value)
                        self.appr = ws['N' + str(x)].value
                        self.bpp = ws['O' + str(x)].value
                        self.cpp = float(ws['P' + str(x)].value)
                        self.awp = ws['Q' + str(x)].value
                        self.bwp = ws['R' + str(x)].value
                        self.cwp = float(ws['S' + str(x)].value)
            # ================= Occupancy factor Τ=================
            if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                if self.d["vraoccup " + str(t)].get() == 1:
                    self.t = float(self.d["occupentry " + str(t)].get())
                elif self.d["vraoccup " + str(t)].get() == 2:
                    ws = wb['Occupancy Factor ( T )']
                    for i in range(2, 31):
                        if self.d["vselocation " + str(t)].get() == ws['A' + str(i)].value:
                            self.t = float(ws['B' + str(i)].value)
            else:
                if self.d["vraoccup " + str(e) + nr].get() == 1:
                    self.t = float(self.d["occupentry " + str(e) + nr].get())
                elif self.d["vraoccup " + str(e) + nr].get() == 2:
                    ws = wb['Occupancy Factor ( T )']
                    for i in range(2, 32):
                        if self.d["vselocation " + str(e) + nr].get() == ws['A' + str(i)].value:
                            self.t = float(ws['B' + str(i)].value)
            # ====================K1" air kerma"================
            ws = wb['unshielding prim']
            self.K1=1
            if self.d["preshuns " + str(e) + nr].get()== 1:
                self.K1 = float(self.d["entk " + str(e) + nr].get())
            else:
                for i in range(1, 6):
                    if self.d["vselroom "+str(t)].get() == "From X-Ray room":
                        if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['C' + str(i)].value)
                    else:
                        if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['C' + str(i)].value)
            #=====================Use Factor=====================
            self.Us = float(self.d["use_ent " + str(e)].get())
            #===============Preshielding=====================
            if self.d["preshvar "+str(e) + nr].get()==1:
                ws = wb["Equiv. thickness of prim pres"]
                if self.d["radiob_pre " + str(e)+nr].get()==1:
                    self.xlead = float(ws['B' + str(3)].value)
                    self.xconc = float(ws['C' + str(3)].value)
                    self.xsteel = float(ws['D' + str(3)].value)
                elif self.d["radiob_pre " + str(e) + nr].get() == 2:
                    self.xlead = float(ws['B' + str(4)].value)
                    self.xconc = float(ws['C' + str(4)].value)
                    self.xsteel = float(ws['D' + str(4)].value)
            else:
                self.xlead = 0
                self.xconc = 0
                self.xsteel = 0
            self.kypol= self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr]
            self.kypol\
                =self.n *self.Us* self.t * self.K1/(self.p ** 2)
            # ===================== Bprimary===============================================================
            self.Bpri = (self.ksec * self.p ** 2) / (self.n *self.Us* self.t * self.K1)
            # ========================== Material selection =================================
            if self.d["vmater "+str(e)+str(o)+nr].get() == "Lead":
                self.thm["xbar " + str(e)+str(o)+nr]  = (1 / (self.alp * self.clp)) * math.log(
                    (self.Bpri ** (-self.clp) + self.blp / self.alp) / (1 + self.blp / self.alp))-self.xlead
                self.xlmat["thic "+str(e) + str(1) + nr] = self.thm["xbar " + str(e)+str(o)+nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Concrete":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.acp * self.ccp)) * math.log(
                    (self.Bpri ** (-self.ccp) + self.bcp / self.acp) / (1 + self.bcp / self.acp))-self.xconc
                self.xlmat["thic " + str(e) + str(2) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Gypsum Wallboard":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.agp * self.cgp)) * math.log(
                    (self.Bpri ** (-self.cgp) + self.bgp / self.agp) / (1 + self.bgp / self.agp))
                self.xlmat["thic " + str(e) + str(3) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Steel":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.asp * self.csp)) * math.log(
                    (self.Bpri ** (-self.csp) + self.bsp / self.asp) / (1 + self.bsp / self.asp))-self.xsteel
                self.xlmat["thic " + str(e) + str(4) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Plate Glass":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.appr * self.cpp)) * math.log(
                    (self.Bpri ** (-self.cpp) + self.bpp / self.appr) / (1 + self.bpp / self.appr))
                self.xlmat["thic " + str(e) + str(5) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Wood":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.awp * self.cwp)) * math.log(
                    (self.Bpri ** (-self.cwp) + self.bwp / self.awp) / (1 + self.bwp / self.awp))
                self.xlmat["thic " + str(e) + str(6) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            if self.d["titleresul " + str(e)+nr] is None:
                self.res["resmat " + str(o)+str(e)]= ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                                text=self.d["vmater " + str(e) + str(o)+nr].get()
                                                                     + ": " + str(
                                                                    round(self.thm["xbar " + str(e)+str(o)+nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o, pady=3, padx=3, sticky="w")
                    op = 0
                elif 2 < o < 5:
                    if o == 3:
                        self.ep += 1
                        op += 1
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o - 2, pady=3, padx=3, sticky="w")
                else:
                    if o == 5:
                        self.ep += 1
                        op += 1
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o - 4, pady=3, padx=3, sticky="w")
            else:

                self.res["resmat " + str(o)+str(e)] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                       text=self.d["vmater " + str(e) + str(o)+nr].get() + ": " + str(
                                                           round(self.thm["xbar " + str(e)+str(o)+nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)]), column=o, pady=3, padx=3,
                                                    sticky="w")
                elif 2 < o < 5:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)] + 1), column=o - 2, pady=3,
                                                    padx=3, sticky="w")
                else:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)] + 2), column=o - 4, pady=3,
                                                    padx=3, sticky="w")

        if self.d["titleresul " + str(e)+nr] is None:
            self.d["titleresul " + str(e)+nr] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                       text=self.barn["lab_bar " + str(e) + nr].cget("text") + ": ")
            self.d["titleresul " + str(e)+nr].grid(row=str(self.ep - op), column=0, pady=3, padx=3, sticky="s")
            self.d["spot {0}".format(str(e))] = self.ep
        else:
            self.d["titleresul " + str(e)+nr].destroy()
            self.d["titleresul " + str(e)+nr] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                       text=self.barn["lab_bar " + str(e) + nr].cget("text") + ": ")
            self.d["titleresul " + str(e)+nr].grid(row=str(self.d["spot " + str(e)]), column=0, pady=3, padx=3,
                                                sticky="s")

        self.rdata={self.barn["lab_bar " + str(e) + nr].cget("text"):{self.d["vmater " + str(e) + str(o)+nr].get():str(
                                                           round(self.thm["xbar " + str(e)+str(o)+nr], 3))}}

        print (self.Bpri)
        self.ep += 1

    def depsecondary(self,e,nr,ep,t):  # Secondary barrier calculations
        if self.d["titleresul " + str(e)+nr] is not None:
            for o in range (1,7):
                if self.res["resmat " + str(o)+str(e)] is not None:
                    self.res["resmat " + str(o)+str(e)].destroy()
        else:
            for o in range(1, 7):
                self.res["resmat {0}".format(str(o)) + str(e)] = None


        for o in range(1, self.d["vnumbmat " + str(e) + nr].get() + 1):
            self.s = float(self.d["entryd " + str(e)+nr].get())
            if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                self.ksec = float(self.d["dikeent " + str(t)].get())
            else:
                self.ksec = float(self.d["dikeent " + str(e) + nr].get())
            from openpyxl import load_workbook  # Opening excel Data
            wb = load_workbook('Data Shielding.xlsx')
            # ========Workload===========
            if self.d["vselroom {0}".format(str(t))].get() == "X-Ray room":
                if self.d["vrawork " + str(t)].get() == 2:
                    self.n = int(self.d["numpapwe " + str(t)].get())
                elif self.d["vrawork " + str(t)].get() == 1:
                    ws = wb['Workload']
                    for x in range(3, 12):
                        if self.d["vsexroom " + str(t)].get() == ws['A' + str(x)].value:
                            self.n = float(self.d["worentry " + str(t)].get())/ (ws['B' + str(x)].value)
                    if self.d["vselxray "+ str(t)].get() == 2:
                        self.n = float(self.d["worentry " + str(t)].get()) /2.5
                ws = wb['sec abc']
                # ====================secondary αβγ============================
                for i in range(3, 18):
                    if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                        self.als = ws['B' + str(i)].value
                        self.bls = ws['C' + str(i)].value
                        self.cls = float(ws['D' + str(i)].value)
                        self.acs = ws['E' + str(i)].value
                        self.bcs = ws['F' + str(i)].value
                        self.ccs = float(ws['G' + str(i)].value)
                        self.ags = ws['H' + str(i)].value
                        self.bgs = ws['I' + str(i)].value
                        self.cgs = float(ws['J' + str(i)].value)
                        self.ass = ws['K' + str(i)].value
                        self.bss = ws['L' + str(i)].value
                        self.css = float(ws['M' + str(i)].value)
                        self.aps = ws['N' + str(i)].value
                        self.bps = ws['O' + str(i)].value
                        self.cps = float(ws['P' + str(i)].value)
                        self.aws = ws['Q' + str(i)].value
                        self.bws = ws['R' + str(i)].value
                        self.cws = float(ws['S' + str(i)].value)

            else:
                if self.d["vrawork " + str(e)+nr].get() == 2:
                    self.n = int(self.d["numpapwe " + str(e)+nr].get())
                elif self.d["vrawork " + str(e)+nr].get() == 1:
                    ws = wb['Workload']
                    for x in range(3, 12):
                        if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(x)].value:
                            self.n = float(self.d["worentry " + str(e)+nr].get())/ (ws['B' + str(x)].value)
                    if self.d["vselxray "+ str(e)+nr].get() == 2:
                        self.n = float(self.d["worentry " + str(e)+nr].get()) /2.5
                ws = wb['sec abc']
                # ====================secondary αβγ============================
                for i in range(3, 18):
                    if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(i)].value:
                        self.als = ws['B' + str(i)].value
                        self.bls = ws['C' + str(i)].value
                        self.cls = float(ws['D' + str(i)].value)
                        self.acs = ws['E' + str(i)].value
                        self.bcs = ws['F' + str(i)].value
                        self.ccs = float(ws['G' + str(i)].value)
                        self.ags = ws['H' + str(i)].value
                        self.bgs = ws['I' + str(i)].value
                        self.cgs = float(ws['J' + str(i)].value)
                        self.ass = ws['K' + str(i)].value
                        self.bss = ws['L' + str(i)].value
                        self.css = float(ws['M' + str(i)].value)
                        self.aps = ws['N' + str(i)].value
                        self.bps = ws['O' + str(i)].value
                        self.cps = float(ws['P' + str(i)].value)
                        self.aws = ws['Q' + str(i)].value
                        self.bws = ws['R' + str(i)].value
                        self.cws = float(ws['S' + str(i)].value)

            # ================= Occupancy factor Τ=================
            if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                if self.d["vraoccup " + str(t)].get() == 1:
                    self.t = float(self.d["occupentry " + str(t)].get())
                elif self.d["vraoccup " + str(t)].get() == 2:
                    ws = wb['Occupancy Factor ( T )']
                    for i in range(2, 31):
                        if self.d["vselocation " + str(t)].get() == ws['A' + str(i)].value:
                            self.t = float(ws['B' + str(i)].value)
            else:
                if self.d["vraoccup " + str(e)+nr].get() == 1:
                    self.t = float(self.d["occupentry " + str(e)+nr].get())
                elif self.d["vraoccup " + str(e)+nr].get() == 2:
                    ws = wb['Occupancy Factor ( T )']
                    for i in range(2, 31):
                        if self.d["vselocation " + str(e)+nr].get() == ws['A' + str(i)].value:
                            self.t = float(ws['B' + str(i)].value)
            #====================K1"secondary air kerma"================

            if self.d["airkerv " + str(e) + nr].get() == 1:
                ws = wb["Uns Air Kerma"]
                if self.d["radiob_leak " + str(e)+nr].get() == 1:
                    for i in range(1,11):
                        if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                            if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(i)].value:
                                self.K1 = float(ws['G' + str(i)].value)
                        else:
                            if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                                self.K1= float(ws['G'+str(i)].value)
                elif self.d["radiob_leak " + str(e)+nr].get() == 2:
                    for i in range(1,11):
                        if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                            if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(i)].value:
                                self.K1= float(ws['I'+str(i)].value)
                        else:
                            if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                                self.K1= float(ws['I'+str(i)].value)
                elif self.d["radiob_leak " + str(e)+nr].get() == 0:
                    for i in range(1,11):
                        if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                            if self.d["vsexroom " + str(e)+nr].get() == ws['A' + str(i)].value:
                                self.K1= float(ws['E'+str(i)].value)
                        else:
                            if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                                self.K1= float(ws['E'+str(i)].value)
            elif self.d["airkerv " + str(e) + nr].get()== 2:
                ws = wb["Uns Air Kerma"]
                for i in range(1, 11):
                    if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                        if self.d["vsexroom " + str(e) + nr].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['F' + str(i)].value)
                    else:
                        if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['F' + str(i)].value)
            elif self.d["airkerv " + str(e) + nr].get() == 3:
                ws = wb["Uns Air Kerma"]
                for i in range(1, 11):
                    if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                        if self.d["vsexroom " + str(e) + nr].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['H' + str(i)].value)
                    else:
                        if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                            self.K1 = float(ws['H' + str(i)].value)
            elif self.d["airkerv " + str(e) + nr].get() == 4:
                self.K1=float(self.d["entk "+ str(e) + nr].get())

            # ===================== Bsec==================================================
            self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr] = self.n * self.t * self.K1/self.s ** 2

            self.Bsec = (self.ksec * self.s ** 2) / (self.n * self.t * self.K1)
            # ========================== Material selection =================================
            if self.d["vmater "+str(e)+str(o)+nr].get() == "Lead":
                self.thm["xbar "+str(e)+str(o)+nr] = (1 / (self.als * self.cls)) * math.log(
                    (self.Bsec ** (-self.cls) + self.bls / self.als) / (1 + self.bls / self.als))
                self.xlmat["thic " + str(e) + str(1) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Concrete":
                self.thm["xbar "+str(e)+str(o)+nr] = (1 / (self.acs * self.ccs)) * math.log(
                    (self.Bsec ** (-self.ccs) + self.bcs / self.acs) / (1 + self.bcs / self.acs))
                self.xlmat["thic " + str(e) + str(2) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Gypsum Wallboard":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.ags * self.cgs)) * math.log(
                    (self.Bsec ** (-self.cgs) + self.bgs / self.ags) / (1 + self.bgs / self.ags))
                self.xlmat["thic " + str(e) + str(3) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Steel":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.ass * self.css)) * math.log(
                    (self.Bsec ** (-self.css) + self.bss / self.ass) / (1 + self.bss / self.ass))
                self.xlmat["thic " + str(e) + str(4) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Plate Glass":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.aps * self.cps)) * math.log(
                    (self.Bsec ** (-self.cps) + self.bps / self.aps) / (1 + self.bps / self.aps))
                self.xlmat["thic " + str(e) + str(5) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Wood":
                self.thm["xbar " + str(e)+str(o)+nr] = (1 / (self.aws * self.cws)) * math.log(
                    (self.Bsec ** (-self.cws) + self.bws / self.aws) / (1 + self.bws / self.aws))
                self.xlmat["thic " + str(e) + str(6) + nr] = self.thm["xbar " + str(e) + str(o) + nr]
            if self.d["titleresul " + str(e)+nr] is None:
                self.res["resmat " + str(o)+str(e)] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                                text=self.d["vmater " + str(e) + str(
                                                                    o)+nr].get() + ": " + str(
                                                                    round(self.thm["xbar " + str(e)+str(o)+nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o, pady=3, padx=3, sticky="w")
                    op=0
                elif 2 < o < 5:
                    if o==3:
                        self.ep+=1
                        op += 1
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o - 2, pady=3, padx=3, sticky="w")
                else:
                    if o==5:
                        self.ep+=1
                        op += 1
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.ep), column=o - 4, pady=3, padx=3, sticky="w")
            else:
                self.res["resmat "+str(o)+str(e)] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                                text=self.d["vmater " + str(e) + str(
                                                                    o)+nr].get() + ": " + str(
                                                                    round(self.thm["xbar " + str(e)+str(o)+nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)]), column=o, pady=3, padx=3, sticky="w")
                elif 2 < o < 5:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)]+1), column=o - 2, pady=3, padx=3, sticky="w")
                else:
                    self.res["resmat " + str(o)+str(e)].grid(row=str(self.d["spot " + str(e)]+2), column=o - 4, pady=3, padx=3, sticky="w")

        if self.d["titleresul " + str(e)+nr] is None:
            self.d["titleresul " + str(e)+nr] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                       text=self.barn["lab_bar " + str(e) + nr].cget("text") + ": ")
            self.d["titleresul " + str(e)+nr].grid(row=str(self.ep-op), column=0, pady=3, padx=3, sticky="s")
            self.d["spot {0}".format(str(e))] = self.ep
        else:
            self.d["titleresul " + str(e)+nr].destroy()
            self.d["titleresul " + str(e)+nr] = ttk.Label(self.d["resultframe " + str(t)+nr], style="AL.TLabel",
                                                       text=self.barn["lab_bar " + str(e) + nr].cget("text") + ": ")
            self.d["titleresul " + str(e)+nr].grid(row=str(self.d["spot " + str(e)]), column=0, pady=3, padx=3,
                                                sticky="s")

        print(self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr])
        self.ep+=1

    def calcom(self,e,nr,b,t):
        self.d["Kco {0}".format(str(t))]=0
        for i in range(1,self.d["vnumbar " +str(t)].get()+1):
            for j in range(1,self.d["vnumwall "+str(t)].get()+1):
                if self.d["vcombar "+nr + str(i)].get() == self.barn["lab_bar " + str(j) + nr].cget("text"):
                    self.d["Kco "+str(t)] += self.d["K "+self.barn["lab_bar " + str(j) + nr].cget("text")+nr]
                    print(self.d["Kco "+str(t)])
                    if self.d["vselroom {0}".format(str(t))].get() == "From X-Ray room":
                        if self.d["dikeent " + str(t)].get() != "":
                            self.k = float(self.d["dikeent " + str(t)].get())
                    else:
                        if self.d["dikeent " + str(j) + nr].get() != "":
                            self.k = float(self.d["dikeent " + str(j) + nr].get())
                    self.d["titleresul " + str(j) + nr].destroy()
                    self.res["resmat " + str(1) + str(j)].destroy()
        self.Bsec = float(self.k)/float(self.d["Kco "+str(t)])

        from openpyxl import load_workbook
        wb = load_workbook('Data Shielding.xlsx')
        if self.d["radiob_w " + str(e) + nr].get() == 2:
            ws = wb['sec abc']
            if self.d["vselroom {0}".format(str(t))].get() == "X-Ray room":
                # ====================secondary αβγ============================
                for i in range(3, 18):
                    if self.d["vsexroom " + str(t)].get() == ws['A' + str(i)].value:
                        self.als = ws['B' + str(i)].value
                        self.bls = ws['C' + str(i)].value
                        self.cls = float(ws['D' + str(i)].value)
                        self.acs = ws['E' + str(i)].value
                        self.bcs = ws['F' + str(i)].value
                        self.ccs = float(ws['G' + str(i)].value)
                        self.ags = ws['H' + str(i)].value
                        self.bgs = ws['I' + str(i)].value
                        self.cgs = float(ws['J' + str(i)].value)
                        self.ass = ws['K' + str(i)].value
                        self.bss = ws['L' + str(i)].value
                        self.css = float(ws['M' + str(i)].value)
                        self.aps = ws['N' + str(i)].value
                        self.bps = ws['O' + str(i)].value
                        self.cps = float(ws['P' + str(i)].value)
                        self.aws = ws['Q' + str(i)].value
                        self.bws = ws['R' + str(i)].value
                        self.cws = float(ws['S' + str(i)].value)
            else:
                for i in range(3, 18):
                    if self.d["vsexroom " + str(e) + nr].get() == ws['A' + str(i)].value:
                        self.als = ws['B' + str(i)].value
                        self.bls = ws['C' + str(i)].value
                        self.cls = float(ws['D' + str(i)].value)
                        self.acs = ws['E' + str(i)].value
                        self.bcs = ws['F' + str(i)].value
                        self.ccs = float(ws['G' + str(i)].value)
                        self.ags = ws['H' + str(i)].value
                        self.bgs = ws['I' + str(i)].value
                        self.cgs = float(ws['J' + str(i)].value)
                        self.ass = ws['K' + str(i)].value
                        self.bss = ws['L' + str(i)].value
                        self.css = float(ws['M' + str(i)].value)
                        self.aps = ws['N' + str(i)].value
                        self.bps = ws['O' + str(i)].value
                        self.cps = float(ws['P' + str(i)].value)
                        self.aws = ws['Q' + str(i)].value
                        self.bws = ws['R' + str(i)].value
                        self.cws = float(ws['S' + str(i)].value)
        elif self.d["radiob_w " + str(e) + nr].get() == 1:
            # ========================primary αβγ========================
            ws = wb['prim abc']
            if self.d["vselroom " + str(t)].get() == "X-Ray room":
                for x in range(2, 39):
                    if self.d["vsexroom " + str(t)].get() == ws['A' + str(x)].value:
                        self.alp = ws['B' + str(x)].value
                        self.blp = ws['C' + str(x)].value
                        self.clp = float(ws['D' + str(x)].value)
                        self.acp = ws['E' + str(x)].value
                        self.bcp = ws['F' + str(x)].value
                        self.ccp = float(ws['G' + str(x)].value)
                        self.agp = ws['H' + str(x)].value
                        self.bgp = ws['I' + str(x)].value
                        self.cgp = float(ws['J' + str(x)].value)
                        self.asp = ws['K' + str(x)].value
                        self.bsp = ws['L' + str(x)].value
                        self.csp = float(ws['M' + str(x)].value)
                        self.appr = ws['N' + str(x)].value
                        self.bpp = ws['O' + str(x)].value
                        self.cpp = float(ws['P' + str(x)].value)
                        self.awp = ws['Q' + str(x)].value
                        self.bwp = ws['R' + str(x)].value
                        self.cwp = float(ws['S' + str(x)].value)
            else:
                for x in range(3, 39):
                    if self.d["vsexroom " + str(e) + nr].get() == ws['A' + str(x)].value:
                        self.alp = ws['B' + str(x)].value
                        self.blp = ws['C' + str(x)].value
                        self.clp = float(ws['D' + str(x)].value)
                        self.acp = ws['E' + str(x)].value
                        self.bcp = ws['F' + str(x)].value
                        self.ccp = float(ws['G' + str(x)].value)
                        self.agp = ws['H' + str(x)].value
                        self.bgp = ws['I' + str(x)].value
                        self.cgp = float(ws['J' + str(x)].value)
                        self.asp = ws['K' + str(x)].value
                        self.bsp = ws['L' + str(x)].value
                        self.csp = float(ws['M' + str(x)].value)
                        self.appr = ws['N' + str(x)].value
                        self.bpp = ws['O' + str(x)].value
                        self.cpp = float(ws['P' + str(x)].value)
                        self.awp = ws['Q' + str(x)].value
                        self.bwp = ws['R' + str(x)].value
                        self.cwp = float(ws['S' + str(x)].value)

        if self.d["vcommater " + str(t)].get() == "Lead":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.als * self.cls)) * math.log(
                (self.Bsec ** (-self.cls) + self.bls / self.als) / (1 + self.bls / self.als))
            self.xlmat["thic {0}".format(str(t))+str(1)] = self.thm["xbar " + str(t) ]
        elif self.d["vcommater " + str(t)].get() == "Concrete":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.acs * self.ccs)) * math.log(
                (self.Bsec ** (-self.ccs) + self.bcs / self.acs) / (1 + self.bcs / self.acs))
            self.xlmat["thic {0}".format(str(t))+str(2)] = self.thm["xbar " + str(t)]
        elif self.d["vcommater " + str(t)].get() == "Gypsum Wallboard":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.ags * self.cgs)) * math.log(
                (self.Bsec ** (-self.cgs) + self.bgs / self.ags) / (1 + self.bgs / self.ags))
            self.xlmat["thic {0}".format(str(t)) + str(3)] = self.thm["xbar " + str(t) ]
        elif self.d["vcommater " + str(t)].get() == "Steel":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.ass * self.css)) * math.log(
                (self.Bsec ** (-self.css) + self.bss / self.ass) / (1 + self.bss / self.ass))
            self.xlmat["thic {0}".format(str(t)) + str(4)] = self.thm["xbar " + str(t) ]
        elif self.d["vcommater " + str(t)].get() == "Plate Glass":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.aps * self.cps)) * math.log(
                (self.Bsec ** (-self.cps) + self.bps / self.aps) / (1 + self.bps / self.aps))
            self.xlmat["thic {0}".format(str(t))+ str(5) ] = self.thm["xbar " + str(t) ]
        elif self.d["vcommater " + str(t)].get() == "Wood":
            self.thm["xbar {0}".format(str(t))] = (1 / (self.aws * self.cws)) * math.log(
                (self.Bsec ** (-self.cws) + self.bws / self.aws) / (1 + self.bws / self.aws))
            self.xlmat["thic {0}".format(str(t)) + str(6)] = self.thm["xbar " + str(t)]

        self.res["coresmat {0}".format(str(t))] = ttk.Label(self.d["resultframe " + str(t) + nr], style="AL.TLabel",
                                text="Combining Barrier: "+self.d["vcommater " + str(t)].get() +
                                ": " + str( round(self.thm["xbar " + str(t)],
                                2)) + " mm")
        self.res["coresmat " + str(t)].grid()
        print(self.thm["xbar "+str(t)])
        self.ep += 1

    def depCTcal(self, e, nr, ep, t):
        if self.d["titleresul " + str(e)+nr] is not None:
            for o in range (1,7):
                if self.res["resmat " + str(o)+str(e)] is not None:
                    self.res["resmat " + str(o)+str(e)].destroy()
        else:
            for o in range(1, 7):
                self.res["resmat {0}".format(str(o)) + str(e)] = None
        for o in range(1, self.d["vnumbmat " + str(e) + nr].get() + 1):
            k1sec_body = float(1.2 * (3 * 10 ** -4) * (1.4 * float(self.d["dlpb_var "+ str(t)].get())))
            k1sec_head = float((9 * 10 ** -5) * (1.4 * float(self.d['dlph_var '+ str(t)].get())))
            self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr] = float((1 / float(self.d["dist_var "+ str(e) + nr].get())) ** 2
                                                                     * ((self.d["bp_var "+str(t)].get() * k1sec_body)
                                                                        + (self.d["hp_var " + str(t)].get() * k1sec_head)))
            B = float(self.d["sh_var " + str(e) + nr].get() / float(self.d["K "+self.barn["lab_bar " + str(e) + nr].cget("text")+nr]))
            print(B)
            if self.d["vmater "+str(e)+str(o)+nr].get() == "Lead":
                if self.d["kvp_var "+ str(t)].get() == 120:
                    a = 2.246
                    b = 5.73
                    c = 0.547
                else:
                    a = 2.009
                    b = 3.99
                    c = 0.342
            elif self.d["vmater "+str(e)+str(o)+nr].get() == "Concrete":
                if self.d["kvp_var "+ str(t)].get() == 120:
                    a = 0.0383
                    b = 0.0142
                    c = 0.658
                else:
                    a = 0.0336
                    b = 0.0122
                    c = 0.519
            self.thm["xbar "+str(e)+str(o)+nr] = float((1 / (a * c)) * math.log((B ** -c + (b / a)) / (1 + (b / a))))

            if self.d["titleresul " + str(e) + nr] is None:
                self.res["resmat " + str(o) + str(e)] = ttk.Label(self.d["resultframe " + str(t) + nr],
                                                                  style="AL.TLabel", text=self.d[
                                                                                              "vmater " + str(e) + str(
                                                                                                  o) + nr].get() + ": " + str(
                        round(self.thm["xbar " + str(e) + str(o) + nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o, pady=3, padx=3, sticky="w")
                    op = 0
                elif 2 < o < 5:
                    if o == 3:
                        self.ep += 1
                        op += 1
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o - 2, pady=3, padx=3,
                                                               sticky="w")
                else:
                    if o == 5:
                        self.ep += 1
                        op += 1
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.ep), column=o - 4, pady=3, padx=3,
                                                               sticky="w")
            else:
                self.res["resmat " + str(o) + str(e)] = ttk.Label(self.d["resultframe " + str(t) + nr],
                                                                  style="AL.TLabel", text=self.d[
                                                                                              "vmater " + str(e) + str(
                                                                                                  o) + nr].get() + ": " + str(
                        round(self.thm["xbar " + str(e) + str(o) + nr], 2)) + " mm")
                if o < 3:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)]), column=o, pady=3,
                                                               padx=3, sticky="w")
                elif 2 < o < 5:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)] + 1), column=o - 2,
                                                               pady=3, padx=3, sticky="w")
                else:
                    self.res["resmat " + str(o) + str(e)].grid(row=str(self.d["spot " + str(e)] + 2), column=o - 4,
                                                               pady=3, padx=3, sticky="w")

        if self.d["titleresul " + str(e) + nr] is None:
            self.d["titleresul " + str(e) + nr] = ttk.Label(self.d["resultframe " + str(t) + nr], style="AL.TLabel",
                                                            text=self.barn["lab_bar " + str(e) + nr].cget(
                                                                "text") + ": ")
            self.d["titleresul " + str(e) + nr].grid(row=str(self.ep - op), column=0, pady=3, padx=3, sticky="s")
            self.d["spot {0}".format(str(e))] = self.ep
        else:
            self.d["titleresul " + str(e) + nr].destroy()
            self.d["titleresul " + str(e) + nr] = ttk.Label(self.d["resultframe " + str(t) + nr], style="AL.TLabel",
                                                            text=self.barn["lab_bar " + str(e) + nr].cget(
                                                                "text") + ": ")
            self.d["titleresul " + str(e) + nr].grid(row=str(self.d["spot " + str(e)]), column=0, pady=3, padx=3,
                                                     sticky="s")

        print(self.d["K " + self.barn["lab_bar " + str(e) + nr].cget("text") + nr])
        self.ep += 1